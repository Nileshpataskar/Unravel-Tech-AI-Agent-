from __future__ import annotations

import os
import re
import sys
import time
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

load_dotenv()

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

SENDER_EMAIL = "nileshcodes04@gmail.com"
SENDER_NAME = "Nilesh Pataskar"
RESUME_PATH = Path(__file__).parent / "Nilesh Pataskar Resume.pdf"
EXCEL_PATH = Path(__file__).parent / "Email Excel.xlsx"

SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587

# Delay between emails (seconds) to avoid Gmail rate limits
DELAY_BETWEEN_EMAILS = 3


# ---------------------------------------------------------------------------
# Cold Email Template
# ---------------------------------------------------------------------------

def build_email_body(contact_name: str, company: str) -> str:
    """Build a personalized cold email body."""
    return f"""\
Hi {contact_name},

I hope this message finds you well.

My name is Nilesh Pataskar, and I'm a Full Stack Developer with roughly three years of experience specializing in TypeScript (React), .NET, SQL, and Python. I'm passionate about building clean, scalable systems and solving real engineering problems.

I came across {company} and I'd love to explore whether there are any open positions — or upcoming opportunities — where my skills could be a good fit.

I've attached my resume for your reference. I'd be grateful for the chance to connect or be considered for any suitable roles.

Thank you for your time, and I look forward to hearing from you.

Warm regards,
Nilesh Pataskar
nileshcodes04@gmail.com
(8888082914)
"""


def build_subject(company: str) -> str:
    """Build a personalized subject line."""
    return f"Exploring Opportunities at {company} — Nilesh Pataskar"


# ---------------------------------------------------------------------------
# Excel Reader
# ---------------------------------------------------------------------------

def extract_email(raw: str) -> str | None:
    """Extract email address from a HYPERLINK formula or plain text."""
    if not raw:
        return None

    raw = str(raw).strip()

    # Match =HYPERLINK("mailto:email","email")
    match = re.search(r'mailto:([^\s",]+)', raw)
    if match:
        return match.group(1).strip().lower()

    # If it looks like a plain email
    match = re.search(r'[\w.+-]+@[\w.-]+\.\w+', raw)
    if match:
        return match.group(0).strip().lower()

    return None


def read_contacts(sheet_index: int = 0) -> list[dict]:
    """Read contacts from Sheet 1 (standard layout: Company, Name, Designation, Phone, Email)."""
    if not EXCEL_PATH.exists():
        print(f"Error: Excel file not found at {EXCEL_PATH}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.worksheets[sheet_index]

    contacts = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        values = [cell.value for cell in row]

        company = str(values[0]).strip() if values[0] else "your company"
        name = str(values[1]).strip() if values[1] else "Hiring Manager"
        email = extract_email(str(values[4])) if values[4] else None

        if email:
            contacts.append({
                "company": company,
                "name": name,
                "email": email,
            })

    return contacts


def read_contacts_sheet4() -> list[dict]:
    """
    Read contacts from Sheet 4 (different layout).
    Each row: Company, Name1, Phone1, Email1, Name2, Phone2, Email2, Name3, Phone3, Email3, ...
    Repeating groups of (Name, Phone, Email) after the company column.
    """
    if not EXCEL_PATH.exists():
        print(f"Error: Excel file not found at {EXCEL_PATH}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.worksheets[5]  # Sheet4 is at index 5

    contacts = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        values = [cell.value for cell in row]

        company = str(values[0]).strip() if values[0] else "your company"

        # Contacts are in groups of 3 starting from index 1: (Name, Phone, Email)
        i = 1
        while i + 2 < len(values):
            name = str(values[i]).strip() if values[i] else None
            email = extract_email(str(values[i + 2])) if values[i + 2] else None

            if name and email:
                contacts.append({
                    "company": company,
                    "name": name,
                    "email": email,
                })
            i += 3  # Move to next (Name, Phone, Email) group

    return contacts


def read_contacts_sheet2(sheet_index: int = 2) -> list[dict]:
    """
    Read contacts from sheets with layout: Sl.No, Client Name, then repeating contact groups.
    Groups are either (Name, Phone, Email) or (Name, Designation, Phone, Email).
    We scan each cell and grab the email wherever we find one, pairing it with
    the most recent name we saw.
    """
    if not EXCEL_PATH.exists():
        print(f"Error: Excel file not found at {EXCEL_PATH}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.worksheets[sheet_index]

    contacts = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        values = [cell.value for cell in row]

        company = str(values[1]).strip() if values[1] else "your company"

        # Scan cells from index 2 onward, tracking last seen name
        last_name = None
        for val in values[2:]:
            if val is None:
                continue
            val_str = str(val).strip()
            email = extract_email(val_str)
            if email:
                contacts.append({
                    "company": company,
                    "name": last_name or "Hiring Manager",
                    "email": email,
                })
                last_name = None  # Reset after pairing
            elif not val_str.isdigit() and '/' not in val_str and len(val_str) > 2 and '@' not in val_str:
                # Likely a name or designation; keep the first one as name
                if last_name is None:
                    last_name = val_str

    return contacts


# ---------------------------------------------------------------------------
# Email Sender
# ---------------------------------------------------------------------------

def send_cold_email(server: smtplib.SMTP, recipient: dict) -> bool:
    """Send a single cold email. Returns True on success."""
    try:
        msg = MIMEMultipart()
        msg["From"] = f"{SENDER_NAME} <{SENDER_EMAIL}>"
        msg["To"] = recipient["email"]
        msg["Subject"] = build_subject(recipient["company"])

        body = build_email_body(recipient["name"], recipient["company"])
        msg.attach(MIMEText(body, "plain"))

        # Attach resume
        with open(RESUME_PATH, "rb") as f:
            attachment = MIMEBase("application", "pdf")
            attachment.set_payload(f.read())
            encoders.encode_base64(attachment)
            attachment.add_header(
                "Content-Disposition",
                f"attachment; filename={RESUME_PATH.name}",
            )
            msg.attach(attachment)

        server.send_message(msg)
        return True
    except Exception as exc:
        print(f"   Failed: {exc}", file=sys.stderr)
        return False


# ---------------------------------------------------------------------------
# CLI entry-point
# ---------------------------------------------------------------------------

def main() -> None:
    """
    Read contacts from Excel, preview, confirm, and send cold emails.

    Usage:
        py bulk_email.py --test             Send one test email to yourself
        py bulk_email.py                     Send to all contacts from Sheet 1
        py bulk_email.py --sheet4            Send to all contacts from Sheet 4
        py bulk_email.py --sheet4 --test     Test with Sheet 4 sample
    """

    test_mode = "--test" in sys.argv
    use_sheet4 = "--sheet4" in sys.argv

    # Validate
    app_password = os.getenv("GMAIL_APP_PASSWORD_CODES04")
    if not app_password:
        print(
            "Error: GMAIL_APP_PASSWORD not set in .env",
            file=sys.stderr,
        )
        sys.exit(1)

    if not RESUME_PATH.exists():
        print(f"Error: Resume not found at {RESUME_PATH}", file=sys.stderr)
        sys.exit(1)

    if test_mode:
        # --- TEST MODE: send one email to yourself ---
        test_contact = {
            "company": "TestCompany Inc.",
            "name": "Nilesh (Test)",
            "email": "nileshcodes04@gmail.com",
        }
        print("TEST MODE: Sending a test email to yourself")
        print(f"  To:      {test_contact['email']}")
        print(f"  Subject: {build_subject(test_contact['company'])}")
        print(f"  Resume:  {RESUME_PATH.name}")
        print()

        confirm = input("Send test email? (y/n): ").strip().lower()
        if confirm != "y":
            print("Cancelled.")
            sys.exit(0)

        print(f"\nConnecting to {SMTP_SERVER}...")
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SENDER_EMAIL, app_password)
            if send_cold_email(server, test_contact):
                print("Test email sent! Check your inbox at nileshcodes04@gmail.com")
            else:
                print("Test email failed.", file=sys.stderr)
        return

    # --- FULL MODE: send to all contacts ---
    contacts = read_contacts_sheet2(sheet_index=4)
    print(f"Found {len(contacts)} contacts in Sheet 10.\n")

    # Preview first 5
    print("Preview (first 5):")
    for i, c in enumerate(contacts[:5], 1):
        print(f"  {i}. {c['name']} ({c['company']}) — {c['email']}")
    if len(contacts) > 5:
        print(f"  ... and {len(contacts) - 5} more\n")

    # Confirmation
    print(f"From:    {SENDER_EMAIL}")
    print(f"Resume:  {RESUME_PATH.name}")
    print(f"Total:   {len(contacts)} emails")
    print()

    confirm = input("Send all emails? (y/n): ").strip().lower()
    if confirm != "y":
        print("Cancelled.")
        sys.exit(0)

    # Send
    sent = 0
    failed = 0

    print(f"\nConnecting to {SMTP_SERVER}...")
    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.starttls()
        server.login(SENDER_EMAIL, app_password)
        print("Connected!\n")

        for i, contact in enumerate(contacts, 1):
            print(f"[{i}/{len(contacts)}] Sending to {contact['name']} ({contact['email']})...", end=" ")
            if send_cold_email(server, contact):
                print("Sent")
                sent += 1
            else:
                print("FAILED")
                failed += 1

            if i < len(contacts):
                time.sleep(DELAY_BETWEEN_EMAILS)

    print(f"\nDone! Sent: {sent} | Failed: {failed}")


if __name__ == "__main__":
    main()
